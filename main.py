import openpyxl  #To edit excel file
import pandas as pd  #To read data
import random  #To randomize
import time  #To delay the process


def display_teams():
    print("GROUP A\n")
    for prnt in range(1, 5):    # Shows all Group B teams
        wb = pd.read_excel("BookA.xlsx", "Sheet" + str(prnt))
        print("\nteam", str(prnt), "\n\n", wb[["Name", "Age", "Batting", "Bowling"]])
    print("\nGROUP B\n")
    for prnt in range(1, 5):    # Shows all Group B teams
        wb = pd.read_excel("BookB.xlsx", "Sheet" + str(prnt))
        print("\nteam", str(prnt), "\n\n", wb[["Name", "Age", "Batting", "Bowling"]])


display_teams()
edit = input("\nDo you need to edit information?\n")

if (edit.lower()) == "yes":
        e_dit = int(input("\nWhich one do you need to edit information?\n1 edit team\n2 edit player\n"))
        e_grp = input("Enter which Group do you want to edit (A or B):\n ")
        e_team = int(input("which team do you want to edit? (1 to 4)\n "))

        if e_dit == 1:
            wb = openpyxl.load_workbook("C:\\Users\\Maneth Naveen\\Desktop\\dfdfd\\doc\\Book" + e_grp + ".xlsx")
            ws = wb["Sheet" + str(e_team)]
            for players in range(1, 12):
                U_name = input("Enter Name of New Player :")
                ws["A" + str(players + 1)].value = U_name
                U_age = input("Enter Age of New Player :")
                ws["B" + str(players + 1)].value = U_age
                U_bat_s = input("Enter Batting Style of New Player :")
                ws["C" + str(players + 1)].value = U_bat_s
                U_bowl_s = input("Enter Bowling Style of New Player :")
                ws["D" + str(players + 1)].value = U_bowl_s
                wb.save("C:\\Users\\Maneth Naveen\\Desktop\\dfdfd\\doc\\Book" + e_grp + ".xlsx")
        elif e_dit == 2:
            while True:
                try:
                    wb = openpyxl.load_workbook("C:\\Users\\Maneth Naveen\\Desktop\\dfdfd\\doc\\Book" + e_grp + ".xlsx")
                    ws = wb["Sheet" + str(e_team)]
                    e_player = int(input("Which player Id? (1 to 11)\n"))
                    e_point = input("Which value you need to update?\n\nA Update Name\nB Update Age\nC Update Batting Style\nD Update Bowling Style\nEnter letter :")
                    e_value = input("Enter the value do you need to input ?\n")
                    ws[(e_point.upper()) + str(e_player + 1)].value = e_value
                    wb.save("C:\\Users\\Maneth Naveen\\Desktop\\dfdfd\\doc\\Book" + e_grp + ".xlsx")
                except (ValueError, FileNotFoundError):
                    print("\n@@@@ Your input is out of index. @@@@\n")
                    continue
                redo = input("Do you need to update anything else?\n")
                if redo == "yes":
                    continue
                else:
                    break
        display_teams()
teams_of_A = []
teams_of_B = []

for grp in range(2):
    if grp == 0:
        G = teams_of_A
        f = 'BookA.xlsx'
    else:
        G = teams_of_B
        f = 'BookB.xlsx'
    for sheet in range(1, 5):
        y = []
        x = pd.read_excel(f, 'Sheet' + str(sheet))
        for i in range(11):
            y.append(x["Name"][i])
        G.append(y)

groupA = {"teams_A1": teams_of_A[0], "teams_A2": teams_of_A[1], "teams_A3": teams_of_A[2], "teams_A4": teams_of_A[3]}
groupB = {"teams_B1": teams_of_B[0], "teams_B2": teams_of_B[1], "teams_B3": teams_of_B[2], "teams_B4": teams_of_B[3]}
groupsA = ["teams_A1", "teams_A2", "teams_A3", "teams_A4"]
groupsB = ["teams_B1", "teams_B2", "teams_B3", "teams_B4"]


def FIRST():
    global bat, ball, fi, si, state
    if f_bb == "First Batting":
        bat = g2[state[a]]
        ball = g2[state[b]]
        fi = state[a]
        si = state[b]
    else:
        ball = g2[state[a]]
        bat = g2[state[b]]
        fi = state[b]
        si = state[a]


def END():
    global  matches,overs,all_B
    if NP == player_state[0]:
        personal_best[player_state[1]] = player1
    elif NP == player_state[1]:
        personal_best[player_state[0]] = player0
    if matches == 0:
        win[fi] = total_score
    elif matches == 1:
        win[si] = total_score
    print("\n       MATCH SUMMERY \n"+str(len(personal_best)) + " players have played.")
    print("They played " + str(overs) + " ove`rs.")
    print(pd.Series(personal_best))
    for oo in list(personal_best):
        if oo in list(all_B):
            x = all_B[oo]+personal_best[oo]
            all_B[oo] = x
        else:
            all_B[oo] = personal_best[oo]

def TOSS():
    global a, b, state, f_bb, bat2, ball2, win
    print("\n" + state[0] + " VS " + state[1])
    toss_team = random.choice(state)
    print("toss team is " + toss_team)
    choose = random.choice(["Head", "Tail"])
    toss = random.choice(["Head", "Tail"])
    f_bb = random.choice(["First Batting", "First Bowling"])
    if choose == toss:
        print(toss_team + " won the toss! They chose " + f_bb + ".")
        if toss_team == state[0]:
            a = 0
            b = 1
            FIRST()
        elif toss_team == state[1]:
            a = 1
            b = 0
            FIRST()
    else:
        print(state[1] + " won the toss! They chose " + f_bb + ".")
        if toss_team == state[0]:
            a = 1
            b = 0
            FIRST()
        elif toss_team == state[1]:
            a = 0
            b = 1
            FIRST()
    bat2 = ball
    ball2 = bat


def WICKET():
    global x, NP, N_bat, o, player_state, player0, player1
    if NP == player_state[0]:
        print(NP + " OUT!! got " + str(player0))
        personal_best[player_state[x]] = player0
        player0 = 0
    else:
        print(NP + " OUT!! got " + str(player1))
        personal_best[player_state[x]] = player1
        player1 = 0
    player_state[x] = N_bat[o]
    NP = player_state[x]
    print("New Player is "+ NP)


def ONE_RUN():
    global NP,player0,player1
    if NP == player_state[0]:
        player0 += 1
        NP = player_state[1]
    else:
        player1 += 1
        NP = player_state[0]


def MATCH():
    global total_score, score, c, NP, player_state, player0, player1, personal_best, x, o, N_bat,matches,overs, all_W, bl
    for matches in range(2):
        if matches == 0:
            N_ball = ball
            N_bat = bat
            wpl = ball
            bl = bat
        elif matches == 1:
            print("\n@@@@@@@@@@@@@@@@@@@\n CHANGING SIDES \n@@@@@@@@@@@@@@@@@@@")
            N_ball = ball2
            N_bat = bat2
            wpl = ball2
            bl = bat2
        scores = [0, 1, 2, "wide ball", 4, 6, "wicket", "no ball"]
        finish_bowling = []
        wickets = []
        in_W = {}
        total_score = 0
        player0 = player1 = 0
        personal_best = {}
        player_state = [N_bat[0], N_bat[1]]
        o = 2
        print("\nOpen batesmen are " + str(player_state[0])+ " and " + str(player_state[1])+".")
        NP = random.choice(player_state)
        try:
            for overs in range(1, 21):
                while True:
                    r_bowler = random.choice(N_ball)
                    if (finish_bowling.count(r_bowler)) == 4:
                        continue
                    else:
                        finish_bowling.append(r_bowler)
                        print("\n\nOVER : " + str(overs) + "   BY :  " + r_bowler)
                        break
                c = 1
                while c <= 6:
                    score = random.choice(scores)
                    if score == 1:
                        total_score += 1
                        print( "Ball: " + str(c) + " --> " + NP + " got 1.")
                        ONE_RUN()
                        c += 1
                    elif score == "wide ball":
                        total_score += 1
                        print("WIDE BALL !!!!")
                    elif score == "no ball":
                        score = random.choice([0, 1, 2, 4, 6])
                        fhw = random.choice(["Hit Wicket", "Run Out", "Catch Out", "LBW"])
                        print(NP +" got "+str(score)+" for NO BALL !!!!")
                        if score == 1:
                            ONE_RUN()
                        else:
                            if NP == player_state[0]:
                                player0 += score
                            else:
                                player1 += score
                        fh = random.choice([0, 1, 2, 4, 6,"wicket"])
                        if fh == 1:
                            ONE_RUN()
                            total_score += (1 + score + fh)
                            print(NP + " got " + str(fh) + " for FREE HIT!!!.")
                        elif fh == "wicket":
                            if fhw =="Run Out":
                                WICKET()
                            else:
                                print(fhw +" !!! "+NP + " NOT OUT!!!")
                        else:
                            if NP == player_state[0]:
                                player0 += fh
                            else:
                                player1 += fh
                            print(NP + " got " + str(fh) + " for FREE HIT!!!.")
                            total_score += (1 + score + fh)
                    elif score == "wicket":
                        print("##### "+random.choice(["Hit Wicket","Run Out","Catch Out","LBW"])+" BY " + r_bowler + " #####")
                        wickets.append(r_bowler)
                        if NP == player_state[0]:
                            x = 0
                            WICKET()
                        elif NP == player_state[1]:
                            x = 1
                            WICKET()
                        c += 1
                        o += 1
                    else:
                        total_score += score
                        print("Ball: " + str(c) + " --> " + NP + " got " + str(score) + ".")
                        if NP == player_state[0]:
                            player0 += score
                        else:
                            player1 += score
                        c += 1
                if NP == player_state[0]:
                    NP = player_state[1]
                elif NP == player_state[1]:
                    NP = player_state[0]
                print("************************************************** \n TOTAL SCORE : " + str(total_score) + "   RUN RATE : " + str(total_score/overs))
                #time.sleep(2)
            END()
        except IndexError:
            END()
        for wp in wpl:
            if wp in wickets:
                in_W[wp] = wickets.count(wp)
                if wp in all_W:
                    tot = (all_W[wp])+(wickets.count(wp))
                    all_W[wp] = tot
                else:
                    all_W[wp] = wickets.count(wp)

all_W = {}
all_B = {}

for a in range(2):
    if a == 0:
        print("\n#################################\n      GROUP A MATCHES\n#################################\n")
        g1 = groupsA
        g2 = groupA
    elif a == 1:
        print("\n#################################\n      GROUP B MATCHES\n#################################\n")
        g1 = groupsB
        g2 = groupB
    for league in range(3):
        print("\n#################################\n      MATCH "+str(league+1)+"\n#################################\n")
        state = random.sample(list(g1), 2)
        win = {}
        TOSS()
        MATCH()
        g1.remove(min(win, key=win.get))
        g2.pop(min(win, key=win.get))
        print("\n*******************************\n   "+max(win, key=win.get)+" won the Match !!!\n*******************************\n")
        print(pd.Series(win))

print("\n#################################\n FINAL LEAGUE HAS BEEN STARTED\n#################################\n")
win={}
state = [groupsA[0], groupsB[0]]
g2 = {groupsA[0]: groupA[state[0]], groupsB[0]: groupB[state[1]]}
TOSS()
MATCH()
state.remove(min(win, key=win.get))
print("\n*******************************\n   "+max(win, key=win.get)+" won the Match !!!\n*******************************\n")
print(pd.Series(win))


print("\n#################################\n        TOP 5 BATSMEN\n#################################\n")
B_df = pd.DataFrame(all_B.items(), columns=['Player', 'Score'])
print((B_df.sort_values(by='Score',ascending=False)).head(5))
print("\n#################################\n        TOP 5 BOWLERS\n#################################\n")
W_df = pd.DataFrame(all_W.items(), columns=['Player', 'Wickets'])
print((W_df.sort_values(by='Wickets',ascending=False)).head(5))